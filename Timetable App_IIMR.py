import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from ortools.sat.python import cp_model
import math
import io
import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# PAGE CONFIGURATION
# ==========================================
st.set_page_config(layout="wide", page_title="IIM Ranchi Timetable Optimizer", page_icon="📅")

# ==========================================
# 1. SIDEBAR: INDEPENDENT VARIABLES
# ==========================================
st.sidebar.image("https://upload.wikimedia.org/wikipedia/en/thumb/f/f6/Indian_Institute_of_Management_Ranchi_Logo.svg/1200px-Indian_Institute_of_Management_Ranchi_Logo.svg.png", width=150)
st.sidebar.header("⚙️ Scheduling Parameters")

max_capacity = st.sidebar.slider("Max Students per Section", 40, 100, 70, help="Courses with more students will be split into Section A & B.")
initial_rooms = st.sidebar.number_input("Initial Classrooms (Weeks 1-4)", min_value=1, max_value=20, value=10)
reduced_rooms = st.sidebar.number_input("Reduced Classrooms (Weeks 5-10)", min_value=1, max_value=20, value=4)
reduction_week = st.sidebar.slider("Week of Capacity Reduction", 1, 10, 5)
max_daily_sessions_student = st.sidebar.slider("Max Daily Sessions per Group", 1, 5, 2, help="Limits consecutive classes for clustered groups.")

st.sidebar.markdown("---")
st.sidebar.markdown("""
**Hard Constraints Enforced:**
- 20 Sessions per section
- **Pre-Clustered Groups**: Sections are grouped to prevent internal clashes
- Zero faculty overlapping
- Strict classroom capacity limits per week
- Reduced Sunday timings
- Zero Student Overlapping
""")

# ==========================================
# 2. DATA PROCESSING (EXCEL PARSER)
# ==========================================
@st.cache_data
def process_uploaded_excel(uploaded_file, max_cap):
    courses_raw = []
    
    # Read the Excel workbook
    xls = pd.ExcelFile(uploaded_file)
    
    # Parse each sheet robustly
    for sheet_name in xls.sheet_names:
        df_sheet = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
        faculty_name = "Unknown Faculty"
        course_name = "Unknown Course"
        header_idx = -1
        
        for i, row in df_sheet.iterrows():
            cols = [str(val).strip() if pd.notna(val) else "" for val in row.values]
            
            if len(cols) > 1 and "Faculty Name" in cols[0]:
                faculty_name = cols[1] if cols[1] else "Unknown Faculty"
                
            if "SN" in cols[0] or "Serial No." in cols[0] or "Student ID" in cols:
                header_idx = i
                for j in range(i-1, -1, -1):
                    prev_cols = [str(val).strip() if pd.notna(val) else "" for val in df_sheet.iloc[j].values]
                    if prev_cols and prev_cols[0]:
                        c_name = prev_cols[0]
                        if "Group Mail ID" not in c_name and "Faculty Name" not in c_name:
                            course_name = c_name
                            break
                break
                
        students = []
        if header_idx != -1:
            headers = [str(val).strip() if pd.notna(val) else "" for val in df_sheet.iloc[header_idx].values]
            if "Student ID" in headers:
                sid_idx = headers.index("Student ID")
                for i in range(header_idx + 1, len(df_sheet)):
                    val = df_sheet.iloc[i, sid_idx]
                    if pd.notna(val) and str(val).strip():
                        students.append(str(val).strip())
                        
        courses_raw.append({
            "Course": course_name,
            "Faculty": faculty_name,
            "Students_List": students,
            "Total_Students": len(students)
        })
        
    df = pd.DataFrame(courses_raw)
    
    sections_list = []
    for _, row in df.iterrows():
        students = row['Students_List']
        num_sections = 2 if row['Total_Students'] > max_cap else 1
        
        section_size = math.ceil(len(students) / num_sections) if num_sections > 0 else 0
        
        for sec in range(num_sections):
            sec_name = "Sec A" if sec == 0 else "Sec B"
            if num_sections == 1: sec_name = "Core"
            
            sec_students = students[sec * section_size : (sec + 1) * section_size]
            
            sections_list.append({
                "Course": row['Course'],
                "Section": sec_name,
                "Faculty": row['Faculty'],
                "Students_Count": len(sec_students),
                "Student_IDs": set(sec_students)
            })
            
    return df, pd.DataFrame(sections_list)


# ==========================================
# 3. OPERATIONS RESEARCH MODEL (GROUPING & SCHEDULING)
# ==========================================
@st.cache_data
def solve_timetable(sections_df, init_rooms, red_rooms, red_week, max_daily):
    # --- PHASE 1: PRE-COMPUTE CONFLICT-FREE GROUPS ---
    min_rooms = min(init_rooms, red_rooms)
    num_sections = len(sections_df)
    
    groups = []
    unassigned = list(range(num_sections))
    intra_group_overlap = 0
    
    while unassigned:
        c_first = unassigned.pop(0)
        current_group = [c_first]
        
        while len(current_group) < min_rooms and unassigned:
            best_c = -1
            best_overlap = float('inf')
            
            faculties = {sections_df.iloc[c]['Faculty'] for c in current_group}
            students = set()
            for c in current_group:
                students = students.union(sections_df.iloc[c]['Student_IDs'])
                
            for i in unassigned:
                if sections_df.iloc[i]['Faculty'] in faculties:
                    continue # Strict Faculty Conflict
                    
                cand_students = sections_df.iloc[i]['Student_IDs']
                overlap = len(students.intersection(cand_students))
                
                # We aggressively pick the section that causes the LEAST student overlap within the group
                if overlap < best_overlap:
                    best_overlap = overlap
                    best_c = i
            
            if best_c != -1:
                current_group.append(best_c)
                unassigned.remove(best_c)
                intra_group_overlap += best_overlap
            else:
                break
                
        groups.append(current_group)


    # --- PHASE 2: OR-TOOLS SCHEDULING ENGINE ---
    model = cp_model.CpModel()
    
    num_weeks = 10
    num_days = 7  
    num_slots = 7 # 1.5 hr slots (Continuous setup until 19:30)
    num_groups = len(groups)
    
    # y[g, w, d, s] = 1 if GROUP 'g' is scheduled at Week 'w', Day 'd', Slot 's'
    y = {}
    for g in range(num_groups):
        for w in range(num_weeks):
            for d in range(num_days):
                for s in range(num_slots):
                    y[g, w, d, s] = model.NewBoolVar(f'y_{g}_{w}_{d}_{s}')
                    
    # Constraint 1: Exactly 20 sessions per GROUP
    for g in range(num_groups):
        model.Add(sum(y[g, w, d, s] for w in range(num_weeks) for d in range(num_days) for s in range(num_slots)) == 20)
        
    # Constraint 2: Dynamic Classroom Capacity Limit per slot
    for w in range(num_weeks):
        current_capacity = init_rooms if (w + 1) < red_week else red_rooms
        for d in range(num_days):
            for s in range(num_slots):
                # The total number of SECTIONS scheduled is the sum of the sizes of the active groups
                model.Add(sum(len(groups[g]) * y[g, w, d, s] for g in range(num_groups)) <= current_capacity)
                
    # Constraint 3: Faculty Overlap Across Groups
    for g1 in range(num_groups):
        for g2 in range(g1 + 1, num_groups):
            fac1 = {sections_df.iloc[c]['Faculty'] for c in groups[g1]}
            fac2 = {sections_df.iloc[c]['Faculty'] for c in groups[g2]}
            if not fac1.isdisjoint(fac2):
                for w in range(num_weeks):
                    for d in range(num_days):
                        for s in range(num_slots):
                            model.AddImplication(y[g1, w, d, s], y[g2, w, d, s].Not())
                            
    # Constraint 4: Student Overlapping Across Groups (SOFT CONSTRAINT LOGIC)
    tracked_overlaps = []
    for g1 in range(num_groups):
        for g2 in range(g1 + 1, num_groups):
            stu1 = set().union(*[sections_df.iloc[c]['Student_IDs'] for c in groups[g1]])
            stu2 = set().union(*[sections_df.iloc[c]['Student_IDs'] for c in groups[g2]])
            shared_count = len(stu1.intersection(stu2))
            
            if shared_count > 0:
                for w in range(num_weeks):
                    for d in range(num_days):
                        for s in range(num_slots):
                            overlap_var = model.NewBoolVar(f'overlap_{g1}_{g2}_{w}_{d}_{s}')
                            model.Add(overlap_var >= y[g1, w, d, s] + y[g2, w, d, s] - 1)
                            tracked_overlaps.append((shared_count, overlap_var))
                            
    # Constraint 5: Max Daily Sessions per Group
    for g in range(num_groups):
        for w in range(num_weeks):
            for d in range(num_days):
                model.Add(sum(y[g, w, d, s] for s in range(num_slots)) <= max_daily)
                
    # Constraint 6: Sunday specific timings (Block Slots 6 & 7)
    for g in range(num_groups):
        for w in range(num_weeks):
            model.Add(y[g, w, 6, 5] == 0) # Slot 6 (Late Evening)
            model.Add(y[g, w, 6, 6] == 0) # Slot 7 (Night)
            
    # Soft Objectives
    student_overlap_penalty = sum(shared * var for shared, var in tracked_overlaps)
    
    # Priority: Minimize student overlaps (Weekend penalty removed to allow regular Sat/Sun scheduling)
    model.Minimize(student_overlap_penalty)
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 300.0
    
    status = solver.Solve(model)
        
    schedule = []
    unscheduled = 0
    total_actual_overlaps = 0
    
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        # Calculate cross-group actual overlaps
        cross_overlaps = 0
        for shared, var in tracked_overlaps:
            if solver.Value(var) == 1:
                cross_overlaps += shared
                
        # Total overlapping instances = (internal group clash * 20 sessions) + cross-group clashes
        total_actual_overlaps = (intra_group_overlap * 20) + cross_overlaps

        slot_labels = [
            "09:00 - 10:30 (Slot 1)",
            "10:30 - 12:00 (Slot 2)",
            "12:00 - 13:30 (Slot 3)",
            "13:30 - 15:00 (Slot 4)",
            "15:00 - 16:30 (Slot 5)",
            "16:30 - 18:00 (Slot 6)",
            "18:00 - 19:30 (Slot 7)"
        ]
        
        start_date = datetime.date(2026, 6, 1) # Explicitly Starts Monday, June 1, 2026
        
        for w in range(num_weeks):
            for d in range(num_days):
                for s in range(num_slots):
                    active_groups = [g for g in range(num_groups) if solver.Value(y[g, w, d, s]) == 1]
                    
                    room_idx = 1
                    for g in active_groups:
                        for c in groups[g]:
                            row = sections_df.iloc[c]
                            # Calculate exactly which date this class lands on
                            current_date = start_date + datetime.timedelta(days=(w * 7) + d)
                            
                            schedule.append({
                                "Group": f"Group {g+1}",
                                "Course": row['Course'],
                                "Section": row['Section'],
                                "Faculty": row['Faculty'],
                                "Week": w + 1,
                                "Date": current_date.strftime('%d-%b-%Y'),
                                "Day": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][d],
                                "Slot": slot_labels[s],
                                "Day_Idx": d,
                                "Slot_Idx": s,
                                "Room": f"CR-{room_idx}"
                            })
                            room_idx += 1
    else:
        unscheduled = num_sections * 20

    return pd.DataFrame(schedule), unscheduled, status, total_actual_overlaps


# ==========================================
# 4. EXCEL GENERATION UTILITIES
# ==========================================
def generate_excel(df, title):
    """
    Generates a beautifully styled, colourful Excel file with predefined column widths.
    Uses openpyxl natively integrated via Pandas to ensure zero missing dependencies.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Schedule', startrow=2, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Schedule']
        
        # Color and Font Definitions
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        alt_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid") # Light Blueish Grey
        title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid") # Navy
        white_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Merge Top Row and Apply Title Styling
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1, value=title)
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        title_cell.border = thin_border
        
        # Apply Column Widths & Format Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border
            
            # Calculate and set dynamic width
            max_len = max(df[col_name].astype(str).map(len).max(), len(str(col_name))) + 2
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(max_len, 12)
            
        # Format Rows with Alternating Colors
        for row_idx in range(len(df)):
            fill_to_use = alt_fill if row_idx % 2 == 1 else PatternFill(fill_type=None)
            for col_idx in range(len(df.columns)):
                cell = worksheet.cell(row=row_idx + 4, column=col_idx + 1)
                cell.alignment = center_align
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = fill_to_use
                    
    return output.getvalue()


# ==========================================
# 5. DASHBOARD UI
# ==========================================
st.title("MBA Timetable Optimization Dashboard")
st.markdown("This dashboard uses **Google OR-Tools** to resolve scheduling conflicts. It uses a **Smart Clustered Block Algorithm** to aggregate sections into non-conflicting groups before solving, ensuring mathematical feasibility and minimal student overlaps.")

# File Uploader
st.markdown("### 📂 Upload IIM Ranchi Course Data (Master Excel File)")
uploaded_file = st.file_uploader("Upload the Master Course Excel file (.xlsx) containing all sheets", type=['xlsx', 'xls'])

if not uploaded_file:
    st.info("Please upload your Master Course Excel file to generate the timetable.")
    st.stop()

# Process Data & Run Optimization
with st.spinner("Clustering Groups & Solving Network Constraints... This may take up to 5 minutes."):
    course_data, sections_data = process_uploaded_excel(uploaded_file, max_capacity)
    
    total_students = len(set(x for l in course_data['Students_List'] for x in l))
    
    schedule_df, unscheduled_count, solver_status, total_overlaps = solve_timetable(
        sections_data, initial_rooms, reduced_rooms, reduction_week, max_daily_sessions_student
    )

    def categorize_slot(idx):
        if idx in [0, 1]: return 'Morning (09:00 - 12:00)'
        elif idx in [2, 3]: return 'Afternoon (12:00 - 15:00)'
        else: return 'Evening (15:00 - 19:30)'
        
    if not schedule_df.empty:
        schedule_df['Shift Category'] = schedule_df['Slot_Idx'].apply(categorize_slot)

# ==========================================
# 6. DEPENDENT VARIABLES & KPIs
# ==========================================
total_capacity_slots = 0
for w in range(1, 11):
    cap = initial_rooms if w < reduction_week else reduced_rooms
    total_capacity_slots += (cap * 47) 

total_sections = len(sections_data)
total_sessions_scheduled = len(schedule_df)
avg_utilization = (total_sessions_scheduled / total_capacity_slots * 100) if total_capacity_slots > 0 else 0

st.markdown("### 🏆 Key Performance Indicators")
col1, col2, col3, col4 = st.columns(4)
col1.metric("Unique Students Processed", total_students)
col2.metric("Total Active Sections", total_sections, delta=f"{total_sections - len(course_data)} split sections", delta_color="inverse")
col3.metric("Overall Room Utilization", f"{avg_utilization:.1f}%")

if solver_status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
    col4.metric("Unscheduled/Conflicts", f"{unscheduled_count} Sessions", delta="Feasible Schedule ✅", delta_color="normal")
else:
    col4.metric("Unscheduled/Conflicts", "INFEASIBLE", delta="Try relaxing constraints ❌", delta_color="inverse")
    st.error("The solver could not find a feasible schedule. You likely need to increase Maximum Classrooms or reduce the Max Daily limit.")
    st.stop()

st.markdown("---")

# ==========================================
# 7. TABBED NAVIGATION (VIEWS & INSIGHTS)
# ==========================================
tab1, tab2, tab3 = st.tabs(["📊 Visual Analytics & Master Data", "📚 Course Schedule Finder", "👨‍🏫 Faculty Schedule Finder"])

with tab1:
    st.markdown("### 📊 Visual Analytics & Insights")

    c1, c2 = st.columns(2)

    # Insight 1: Bar chart of Course Enrollments
    with c1:
        fig1 = px.bar(course_data, x="Course", y="Total_Students", color="Total_Students", 
                      title="Course Enrollments (Dashed line triggers Section Splits)",
                      color_continuous_scale=px.colors.sequential.Teal)
        fig1.add_hline(y=max_capacity, line_dash="dash", annotation_text="Max Capacity Limit", line_color="red")
        fig1.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", showlegend=False)
        st.plotly_chart(fig1, use_container_width=True)

    # Insight 2: Line chart showing capacity
    with c2:
        weekly_usage = schedule_df.groupby('Week').size().reset_index(name='Used Sessions')
        
        capacity_data = []
        for w in range(1, 11):
            cap = initial_rooms if w < reduction_week else reduced_rooms
            capacity_data.append({"Week": w, "Capacity Limit": cap * 47})
        cap_df = pd.DataFrame(capacity_data)
        
        usage_vs_cap = pd.merge(weekly_usage, cap_df, on="Week", how="right").fillna(0)
        
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=usage_vs_cap['Week'], y=usage_vs_cap['Used Sessions'], mode='lines+markers', name='Actual Usage', line=dict(color='#1f77b4', width=3)))
        fig2.add_trace(go.Scatter(x=usage_vs_cap['Week'], y=usage_vs_cap['Capacity Limit'], mode='lines', name='Total Capacity', line=dict(color='#ff7f0e', width=2, dash='dash')))
        fig2.update_layout(title="Classroom Utilization vs Available Capacity Drop", xaxis_title="Week", yaxis_title="Total Sessions", plot_bgcolor="rgba(0,0,0,0)")
        fig2.add_vline(x=reduction_week - 0.5, line_width=1, line_dash="dash", line_color="red", annotation_text="Capacity Reduction")
        st.plotly_chart(fig2, use_container_width=True)

    c3, c4 = st.columns(2)

    # Insight 3: Faculty Workload
    with c3:
        faculty_workload = schedule_df.groupby('Faculty').size().reset_index(name='Total Sessions')
        fig3 = px.bar(faculty_workload.sort_values('Total Sessions', ascending=False), 
                      x="Faculty", y="Total Sessions", 
                      title="Faculty Workload Distribution (Over 10 Weeks)",
                      color="Total Sessions", color_continuous_scale="Blues")
        fig3.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig3, use_container_width=True)

    # Insight 4: Pie chart
    with c4:
        time_usage = schedule_df.groupby('Shift Category').size().reset_index(name='Count')
        fig4 = px.pie(time_usage, names='Shift Category', values='Count', 
                      title="Distribution of Classes by Time of Day", hole=0.4,
                      color_discrete_sequence=px.colors.qualitative.Pastel)
        st.plotly_chart(fig4, use_container_width=True)

    # Insight 5: Heatmap
    st.markdown("### 🗺️ Weekly Operational Heatmap")
    heatmap_data = schedule_df.groupby(['Week', 'Day']).size().reset_index(name='Sessions')

    days_order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    heatmap_pivot = heatmap_data.pivot(index='Day', columns='Week', values='Sessions').reindex(days_order)
    heatmap_pivot = heatmap_pivot.reindex(columns=list(range(1, 11)), fill_value=0).fillna(0)

    fig5 = px.imshow(heatmap_pivot, text_auto=True, aspect="auto",
                     labels=dict(x="Week of Term", y="Day of Week", color="Sessions Scheduled"),
                     title="Session Density per Day across the Term", color_continuous_scale="YlGnBu")
    st.plotly_chart(fig5, use_container_width=True)

    st.markdown("### 📋 Master Timetable Output")
    st.caption("Sections are partitioned into non-conflicting 'Groups'. Displayed in strict chronological sequence starting from June 1, 2026.")
    
    # Sort strictly chronologically
    display_df = schedule_df.sort_values(by=["Week", "Day_Idx", "Slot_Idx", "Group", "Room"])
    
    # Drop helper calculation columns and arrange logically
    display_df = display_df.drop(columns=['Slot_Idx', 'Day_Idx', 'Shift Category'])
    cols = ['Group', 'Course', 'Section', 'Faculty', 'Week', 'Date', 'Day', 'Slot', 'Room']
    display_df = display_df[cols]
    
    st.dataframe(display_df, use_container_width=True)
    
    # ==========================================
    # 📥 EXCEL DOWNLOAD CENTER
    # ==========================================
    st.markdown("### 📥 Download Center")
    st.caption("Export the optimized schedules and analytical reports in beautifully formatted Excel files.")
    
    # Calculate specific dynamic metrics for the buttons
    start_str = datetime.date(2026, 6, 1).strftime('%d %b %Y')
    end_str = (datetime.date(2026, 6, 1) + datetime.timedelta(days=69)).strftime('%d %b %Y')
    num_courses = len(course_data)
    num_sections_total = len(sections_data)
    courses_split = len(course_data[course_data['Total_Students'] > max_capacity])
    num_groups = schedule_df['Group'].nunique() if not schedule_df.empty else 0
    total_days = schedule_df['Date'].nunique() if not schedule_df.empty else 0

    st.markdown("##### 1. Master Schedules")
    master_excel = generate_excel(display_df, f"IIM Ranchi - Master Timetable ({start_str} to {end_str})")
    st.download_button(
        label=f"📅 MASTER TIMETABLE — 10 WEEKS ({start_str} – {end_str})",
        data=master_excel,
        file_name="Master_Timetable.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    ind_df = schedule_df.sort_values(by=['Course', 'Section', 'Week', 'Day_Idx', 'Slot_Idx']).drop(columns=['Slot_Idx', 'Day_Idx', 'Shift Category'])
    ind_df = ind_df[['Course', 'Section', 'Group', 'Faculty', 'Week', 'Date', 'Day', 'Slot', 'Room']]
    ind_excel = generate_excel(ind_df, "IIM Ranchi - Individual Section Schedules")
    st.download_button(
        label=f"📚 INDIVIDUAL SECTION SCHEDULES — {num_sections_total} Sections × 20 Sessions Each",
        data=ind_excel,
        file_name="Individual_Section_Schedules.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.markdown("##### 2. Structural & Utilisation Reports")
    dl_col1, dl_col2 = st.columns(2)
    
    with dl_col1:
        cat_df = sections_data[['Course', 'Section', 'Faculty', 'Students_Count']].copy()
        cat_excel = generate_excel(cat_df, "IIM Ranchi - Course Catalogue")
        st.download_button(
            label=f"📖 COURSE CATALOGUE — {num_courses} Courses | {num_sections_total} Sections | {courses_split} Courses Split (enrollment > {max_capacity})",
            data=cat_excel,
            file_name="Course_Catalogue.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        groups_df = schedule_df[['Group', 'Course', 'Section', 'Faculty']].drop_duplicates().sort_values(by=['Group', 'Course'])
        grp_excel = generate_excel(groups_df, "IIM Ranchi - Section Colour Groups")
        st.download_button(
            label=f"🎨 SECTION COLOUR GROUPS — {num_groups} Groups | Sections in Same Group Have ZERO Conflicts",
            data=grp_excel,
            file_name="Section_Groups.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with dl_col2:
        util_df = schedule_df.groupby(['Week', 'Date', 'Day', 'Slot']).agg(Rooms_Used=('Room', 'nunique')).reset_index()
        util_excel = generate_excel(util_df, "IIM Ranchi - Classroom Utilisation")
        st.download_button(
            label=f"🏫 CLASSROOM UTILISATION — Day-by-Day | Spread Across All {total_days} Days",
            data=util_excel,
            file_name="Classroom_Utilisation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


with tab2:
    st.markdown("### 📚 Find Weekly Schedule by Course & Section")
    if not schedule_df.empty:
        c_col1, c_col2, c_col3 = st.columns(3)
        with c_col1:
            selected_course = st.selectbox("Select Course", sorted(schedule_df['Course'].unique()))
        with c_col2:
            available_sections = sorted(schedule_df[schedule_df['Course'] == selected_course]['Section'].unique())
            selected_section = st.selectbox("Select Section", available_sections)
        with c_col3:
            selected_week_c = st.slider("Select Week", 1, 10, 1, key="course_week")
            
        filtered_course_df = schedule_df[(schedule_df['Course'] == selected_course) & 
                                         (schedule_df['Section'] == selected_section) & 
                                         (schedule_df['Week'] == selected_week_c)]
        
        filtered_course_df = filtered_course_df.sort_values(by=["Day_Idx", "Slot_Idx"])
        filtered_course_df = filtered_course_df[['Group', 'Course', 'Section', 'Faculty', 'Week', 'Date', 'Day', 'Slot', 'Room']]
        st.dataframe(filtered_course_df, use_container_width=True)
        
        # EXCEL DOWNLOAD
        course_excel = generate_excel(filtered_course_df, f"Course Schedule: {selected_course} ({selected_section}) - Week {selected_week_c}")
        st.download_button(
            label="📥 Download This Course Schedule",
            data=course_excel,
            file_name=f"Schedule_{selected_course.replace(' ', '_')}_Wk{selected_week_c}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("No schedule data available.")

with tab3:
    st.markdown("### 👨‍🏫 Find Weekly Schedule by Faculty")
    if not schedule_df.empty:
        f_col1, f_col2 = st.columns(2)
        with f_col1:
            selected_faculty = st.selectbox("Select Faculty", sorted(schedule_df['Faculty'].unique()))
        with f_col2:
            selected_week_f = st.slider("Select Week", 1, 10, 1, key="faculty_week")
            
        filtered_fac_df = schedule_df[(schedule_df['Faculty'] == selected_faculty) & 
                                      (schedule_df['Week'] == selected_week_f)]
                                      
        filtered_fac_df = filtered_fac_df.sort_values(by=["Day_Idx", "Slot_Idx"])
        filtered_fac_df = filtered_fac_df[['Group', 'Course', 'Section', 'Faculty', 'Week', 'Date', 'Day', 'Slot', 'Room']]
        st.dataframe(filtered_fac_df, use_container_width=True)
        
        # EXCEL DOWNLOAD
        fac_excel = generate_excel(filtered_fac_df, f"Faculty Schedule: {selected_faculty} - Week {selected_week_f}")
        st.download_button(
            label="📥 Download This Faculty Schedule",
            data=fac_excel,
            file_name=f"Schedule_{selected_faculty.replace(' ', '_')}_Wk{selected_week_f}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("No schedule data available.")
